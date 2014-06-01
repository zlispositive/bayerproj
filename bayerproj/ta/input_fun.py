from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
from ta.models import Business_Unit, Testing_Machine, Accessory, Chemical, Testing_Activity


import sys
reload(sys)
sys.setdefaultencoding('utf-8')

def input_fun():
	wb = load_workbook('/Users/zhangling/Dropbox/bayerproj/inputinfo/inputinfo.xlsx')
	# wb = load_workbook('empty_book.xlsx')

	s= '''
	+-------------------------------------------+
	|                                           |
	|                                           |
	|                                           |
	|             New Work Sheet                |
	|                                           |
	|                                           |
	|                                           |
	+-------------------------------------------+

		'''
	# print s

	for ws in wb:
		print ws.title
		highest_column = ws.get_highest_column()
		highest_row = ws.get_highest_row()

		print highest_row, highest_column
		# for col_index in xrange(1, 10):
		# 	col = get_colum_letter(col_index)
		
		# for col_index in xrange(1, highest_column+1):
		# 	col = get_column_letter(col_index)
		# 	for row in xrange(1, highest_row+1):
		# 		# print ws.cell('%s%s' % (col, row)).value
		# 		print '%s%s' % (col, row)

		for row in xrange(2, highest_row+1):
			for col_index in xrange(1, highest_column+1):
				col = get_column_letter(col_index)
				# print '%s%s' % (col, row)
				# print '%s%s: %s' % (col, row, ws.cell('%s%s' % (col, row)).value)
				print '%s: %s' % (ws.cell('%s%s' % (col, 1)).value, ws.cell('%s%s' % (col, row)).value)
				# print '%s  %s' % ('p'+('%d' % row), ('%s' % ws.title))

def Business_Unit_input_fun():
	wb = load_workbook('/Users/zhangling/Dropbox/bayerproj/inputinfo/inputinfo.xlsx')
	ws = wb['Business_Unit']
	# print ws.title
	highest_column = ws.get_highest_column()
	highest_row = ws.get_highest_row()

	for row in xrange(2, highest_row+1):
		dic1 = {}
		for col_index in xrange(1, highest_column+1):
			col = get_column_letter(col_index)
			print '%s: %s' % (ws.cell('%s%s' % (col, 1)).value, ws.cell('%s%s' % (col, row)).value)
			dic1['%s' % ws.cell('%s%s' % (col, 1)).value] = str(ws.cell('%s%s' % (col, row)).value)
		print dic1
		p = Business_Unit.objects.create(bu_name=dic1['bu_name'], bu_staff=dic1['bu_staff'], email=dic1['email'])
		p.save()
	Business_Unit.objects.order_by("bu_staff")

def Machine_input_fun():
	wb = load_workbook('/Users/zhangling/Dropbox/bayerproj/inputinfo/inputinfo.xlsx')
	ws = wb['Testing_Machine']
	# print ws.title
	highest_column = ws.get_highest_column()
	highest_row = ws.get_highest_row()

	for row in xrange(2, highest_row+1):
		dic1 = {}
		for col_index in xrange(1, highest_column+1):
			col = get_column_letter(col_index)
			print '%s: %s' % (ws.cell('%s%s' % (col, 1)).value, ws.cell('%s%s' % (col, row)).value)
			dic1['%s' % ws.cell('%s%s' % (col, 1)).value] = str(ws.cell('%s%s' % (col, row)).value)
		print dic1
		p = Testing_Machine.objects.create(owner=dic1['owner'], status=dic1['status'], equipment_code=dic1['equipment_code'], equipment_name=dic1['equipment_name'], equipment_model=dic1['equipment_model'], equipment_amount=dic1['equipment_amount'], equipment_shared=dic1['equipment_shared'], capacity_theoretical=dic1['capacity_theoretical'], plained_downtime=dic1['plained_downtime'], plained_nonwork=dic1['plained_nonwork'], capacity_practical=dic1['capacity_practical'], adjustment_ratio=dic1['adjustment_ratio'], capacity_normal=dic1['capacity_normal'], waiting_time=dic1['waiting_time'], processing_time=dic1['processing_time'])
		p.save()

def Accessory_input_fun():
	wb = load_workbook('/Users/zhangling/Dropbox/bayerproj/inputinfo/inputinfo.xlsx')
	ws = wb['Accessory']
	# print ws.title
	highest_column = ws.get_highest_column()
	highest_row = ws.get_highest_row()

	for row in xrange(2, highest_row+1):
		dic1 = {}
		for col_index in xrange(1, highest_column+1):
			col = get_column_letter(col_index)
			print '%s: %s' % (ws.cell('%s%s' % (col, 1)).value, ws.cell('%s%s' % (col, row)).value)
			dic1['%s' % ws.cell('%s%s' % (col, 1)).value] = str(ws.cell('%s%s' % (col, row)).value)
		print dic1
		p = Accessory.objects.create(accessory_item=dic1['accessory_item'], accessory_price=dic1['accessory_price'], accessory_reusalbe=dic1['accessory_reusalbe'])
		p.save()

def Chemical_input_fun():
	wb = load_workbook('/Users/zhangling/Dropbox/bayerproj/inputinfo/inputinfo.xlsx')
	ws = wb['Chemical']
	# print ws.title
	highest_column = ws.get_highest_column()
	highest_row = ws.get_highest_row()

	for row in xrange(2, highest_row+1):
		dic1 = {}
		for col_index in xrange(1, highest_column+1):
			col = get_column_letter(col_index)
			print '%s: %s' % (ws.cell('%s%s' % (col, 1)).value, ws.cell('%s%s' % (col, row)).value)
			dic1['%s' % ws.cell('%s%s' % (col, 1)).value] = str(ws.cell('%s%s' % (col, row)).value)
		print dic1
		p = Chemical.objects.create(chem_item=dic1['chem_item'], chem_unit_cost=dic1['chem_unit_cost'])
		p.save()

def Activity_input_fun():
	wb = load_workbook('/Users/zhangling/Dropbox/bayerproj/inputinfo/inputinfo.xlsx')
	ws = wb['Testing_Activity']
	# print ws.title
	highest_column = ws.get_highest_column()
	highest_row = ws.get_highest_row()

	for row in xrange(2, highest_row+1):
		dic1 = {}
		for col_index in xrange(1, highest_column+1):
			col = get_column_letter(col_index)
			print '%s: %s' % (ws.cell('%s%s' % (col, 1)).value, ws.cell('%s%s' % (col, row)).value)
			dic1['%s' % ws.cell('%s%s' % (col, 1)).value] = str(ws.cell('%s%s' % (col, row)).value)
		print dic1
		p = Testing_Activity.objects.create(bu=dic1['bu'], testing_type=dic1['testing_type'], testing_name=dic1['testing_name'], testing_material=dic1['testing_material'], testing_capability1=['testing_capability1'], testing_capability2=['testing_capability2'], time_ad=['time_ad'], setuptime_machine=['setuptime_machine'], setuptime_labor=['setuptime_labor'], runtime_machine=['runtime_machine'], runtime_labor=['runtime_labor'], conditiontime=['conditiontime'], outsourcing_price=['outsourcing_price'], outsourcing_partner=['outsourcing_partner'], outsourcing_risk=['outsourcing_risk'], outsourcing_cost=['outsourcing_cost'])
		p.save()

def datainput():
	Business_Unit_input_fun()
	Machine_input_fun()
	Accessory_input_fun()
	Chemical_input_fun()
	Activity_input_fun()