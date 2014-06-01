from openpyxl import load_workbook
from openpyxl.cell import get_column_letter

def input_fun():
	wb = load_workbook('/Users/zhangling/Dropbox/bayerproj/inputinfo/inputinfo.xlsx')
	# wb = load_workbook('empty_book.xlsx')

	s= '''
	+-------------------------------------------+
	|                                           |
	|                                           |
	|                                           |
	|             New Work Sheet                |
	|                                           |
	|                                           |
	|                                           |
	+-------------------------------------------+

		'''
	# print s

	for ws in wb:
		print ws.title
		highest_column = ws.get_highest_column()
		highest_row = ws.get_highest_row()

		print highest_row, highest_column
		# for col_index in xrange(1, 10):
		# 	col = get_colum_letter(col_index)
		
		# for col_index in xrange(1, highest_column+1):
		# 	col = get_column_letter(col_index)
		# 	for row in xrange(1, highest_row+1):
		# 		# print ws.cell('%s%s' % (col, row)).value
		# 		print '%s%s' % (col, row)

		for row in xrange(2, highest_row+1):
			for col_index in xrange(1, highest_column+1):
				col = get_column_letter(col_index)
				# print '%s%s' % (col, row)
				# print '%s%s: %s' % (col, row, ws.cell('%s%s' % (col, row)).value)
				print '%s: %s' % (ws.cell('%s%s' % (col, 1)).value, ws.cell('%s%s' % (col, row)).value)
				# print '%s  %s' % ('p'+('%d' % row), ('%s' % ws.title))

def input_fun_Business_Unit():
	wb = load_workbook('/Users/zhangling/Dropbox/bayerproj/inputinfo/inputinfo.xlsx')
	ws = wb['Business_Unit']
	# print ws.title
	highest_column = ws.get_highest_column()
	highest_row = ws.get_highest_row()

	for row in xrange(2, highest_row+1):
		dic1 = {}
		for col_index in xrange(1, highest_column+1):
			col = get_column_letter(col_index)
			print '%s: %s' % (ws.cell('%s%s' % (col, 1)).value, ws.cell('%s%s' % (col, row)).value)
			dic1['%s' % ws.cell('%s%s' % (col, 1)).value] = str(ws.cell('%s%s' % (col, row)).value)
		print dic1
		p = Business_Unit.objects.create(bu_name=dic1['bu_name'], bu_staff=dic1['bu_staff'], email=dic1['email'])
		p.save()

# input_fun()
input_fun_Business_Unit()