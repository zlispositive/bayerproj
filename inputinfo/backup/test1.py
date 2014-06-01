from openpyxl import Workbook

from openpyxl.cell import get_column_letter

wb = Workbook()

dest_filename = r'empty_book.xlsx'

ws = wb.active

ws.title = "range names"

for col_idx in xrange(1, 4):
    col = get_column_letter(col_idx)
    for row in xrange(1, 6):
    	print '%s%s' % (col, row)
        ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)

ws = wb.create_sheet()

ws.title = 'Pi'

ws['F5'] = 3.14

wb.save(filename = dest_filename)